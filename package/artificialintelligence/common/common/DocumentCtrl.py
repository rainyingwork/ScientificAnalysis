import openpyxl
from openpyxl.comments import Comment
from package.artificialintelligence.common.common.AnalysisFunction import AnalysisFunction

class DocumentCtrl:

    def __init__(self):
        pass

    def makeAnalysisDoc(self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = AnalysisFunction.getDataColumnNameArr()

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(initFilePath)

        for key in dataMap.keys():
            columnNumber = 4
            detailDataMapArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for detailDataMap in detailDataMapArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["Project"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["TableName"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["Memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["StartDate"] if "StartDate" in detailDataMap["TableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["EndDate"] if "EndDate" in detailDataMap["TableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["DataType"]))
                ws.cell(row=8, column=columnNumber, value="{}".format(detailDataMap["product"]["description"]))
                ws.cell(row=9, column=columnNumber, value="{}".format(detailDataMap["project"]["description"]))
                ws.cell(row=10, column=columnNumber, value="{}".format(detailDataMap["version"]["description"]))
                ws.cell(row=11, column=columnNumber, value="{}".format(detailDataMap["dt"]["description"]))
                ws.cell(row=12, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["DataType"]))

                if "CommentMemo" in detailDataMap["TableInfo"].keys() :
                    commentMemoComment = Comment(text="{}".format(detailDataMap["TableInfo"]["CommentMemo"]), author="Code", height=100, width=500)
                    ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(4))].comment = commentMemoComment

                rowNumber = 12
                for dataColumn in dataColumnArr:
                    rowNumber = rowNumber + 1
                    if dataColumn != detailDataMap[dataColumn]["description"]:
                        description = detailDataMap[dataColumn]["description"] if "description" in detailDataMap[dataColumn].keys() else ""
                        memo = detailDataMap[dataColumn]["memo"] if "memo" in detailDataMap[dataColumn].keys() else ""
                        commentMemo = ""
                        processfunc = ",".join(detailDataMap[dataColumn]["processfuncs"]) if "processfuncs" in detailDataMap[dataColumn].keys() else ""
                        checkfunc = ",".join(detailDataMap[dataColumn]["checkfuncs"]) if "checkfuncs" in detailDataMap[dataColumn].keys() else ""

                        if dataColumn == detailDataMap[dataColumn]["description"]:
                            continue

                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))

                        if memo != "" or commentMemo != "" or checkfunc != "" :
                            ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = Comment(
                                text="備註: {}\n其他備註: {}\n處理方式: {}\n檢查方式: {}".format(memo, commentMemo,processfunc,checkfunc)
                                , author="Code",height=150, width=800
                            )

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)



    def makeAnalysisDoubleInfoByDataBase(self,product,project,version,dt) :
        import os , json
        from dotenv import load_dotenv
        from package.common.common.database.PostgresCtrl import PostgresCtrl
        load_dotenv(dotenv_path="env/postgresql.env")
        postgresCtrl = PostgresCtrl(
            host=os.getenv("POSTGRES_HOST")
            , port=int(os.getenv("POSTGRES_POST"))
            , user=os.getenv("POSTGRES_USERNAME")
            , password=os.getenv("POSTGRES_PASSWORD")
            , database=os.environ["POSTGRES_OPSNABAGEMENT_DATABASE"]
            , schema=os.environ["POSTGRES_OPSNABAGEMENT_SCHEMA"]
        )

        sqlReplaceArr = [
            ["[:Product]",product],
            ["[:Project]",project],
            ["[:Version]",version],
            ["[:DataNoLine]",dt],
        ]

        searchSQL = """
            select 
                common_011	
                , common_012	
                , common_013	
                , common_014	
                , common_015
            from observationdata.analysisdata AA
            where 1 = 1 
                and AA.product = '[:Product]'
                and AA.project = '[:Project]'
                and AA.version = '[:Version]'
                and AA.dt = '[:DataNoLine]' ; 
        """
        for sqlReplace in sqlReplaceArr :
            searchSQL = searchSQL.replace(sqlReplace[0],sqlReplace[1])

        analysisDoubleInfoMap = {}
        analysisDoubleInfoDF = postgresCtrl.searchSQL(searchSQL)
        for index , row in analysisDoubleInfoDF.iterrows() :
            columnName = "double_" + str(row['common_013']).zfill(3)
            columnDetailInfo =json.loads(row['common_015'])
            analysisDoubleInfoMap[columnName] ={
                "description" : row['common_012'] ,
                'datatype': 'double',
                "memo" : row['common_011'] + ',' + row['common_014'] ,
                "processfuncs":columnDetailInfo["DataPreProcess"]["ProcessingOrder"] if "DataPreProcess" in columnDetailInfo.keys() else [] ,
                "checkfuncs":columnDetailInfo["DataCheck"]["CheckFunction"] if "DataCheck" in columnDetailInfo.keys() else [] ,
            }

        return analysisDoubleInfoMap