import openpyxl
from openpyxl.comments import Comment
from package.dataengineer.common.standard.StandardFunction import StandardFunction

class DocumentCtrl:

    def __init__(self):
        pass

    def MakeStandardDoc(self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = StandardFunction.getDataColumnNameArr()

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(initFilePath)

        for key in dataMap.keys():
            columnNumber = 4
            detailDataMapArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for detailDataMap in detailDataMapArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["Project"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["TableName"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["Memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["StartDate"] if "StartDate" in detailDataMap["TableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["EndDate"] if "EndDate" in detailDataMap["TableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["DataType"]))
                ws.cell(row=8, column=columnNumber, value="{}".format(detailDataMap["product"]["description"]))
                ws.cell(row=9, column=columnNumber, value="{}".format(detailDataMap["project"]["description"]))
                ws.cell(row=10, column=columnNumber, value="{}".format(detailDataMap["tablename"]["description"]))
                ws.cell(row=11, column=columnNumber, value="{}".format(detailDataMap["dt"]["description"]))
                ws.cell(row=12, column=columnNumber, value="{}".format(detailDataMap["TableInfo"]["DataType"]))

                if "CommentMemo" in detailDataMap["TableInfo"].keys() :
                    commentMemoComment = Comment(text="{}".format(detailDataMap["TableInfo"]["CommentMemo"]), author="Code", height=100, width=500)
                    ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(4))].comment = commentMemoComment

                rowNumber = 12
                for dataColumn in dataColumnArr:
                    rowNumber = rowNumber + 1
                    description = detailDataMap[dataColumn]["description"] if "description" in detailDataMap[dataColumn].keys() else ""
                    memo = detailDataMap[dataColumn]["memo"] if "memo" in detailDataMap[dataColumn].keys() else ""
                    commentMemo = detailDataMap[dataColumn]["commentMemo"] if "commentMemo" in detailDataMap[dataColumn].keys() else ""
                    if dataColumn == detailDataMap[dataColumn]["description"]:
                        continue
                    ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))
                    if memo == "" and commentMemo == "" :
                        continue
                    columnCommentMemoComment = Comment(text="備註: {}\n其他備註: {}".format(memo, commentMemo), author="Code", height=100, width=500)
                    ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = columnCommentMemoComment

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)
